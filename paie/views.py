from multiprocessing import context
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from datetime import datetime, timedelta
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.shortcuts import render, redirect
from django.contrib import messages
from core.models import Salaire
from paie.services.salaire_liste_service import SalaireListeService
from paie.services.salaire_generation_service import SalaireGenerationService

import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse

from calendar import monthrange
from datetime import date
from django.utils import timezone


from calendar import monthrange
from datetime import date
from django.utils import timezone

class SalaireLectureView(LoginRequiredMixin, TemplateView):
    template_name = "paie/salaire_liste.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        today = timezone.now().date()

        # ----------------------------
        # MOIS / ANNÉE
        # ----------------------------
        month = self.request.GET.get("month")
        year = self.request.GET.get("year")

        if month and year:
            month = int(month)
            year = int(year)
        else:
            month = today.month
            year = today.year

        last_day = monthrange(year, month)[1]
        date_debut = date(year, month, 1)
        date_fin = date(year, month, last_day)

        # ----------------------------
        # SERVICE SALAIRES
        # ----------------------------
        result = SalaireListeService.get_salaires(
            date_debut=date_debut,
            date_fin=date_fin,
        )

        # ----------------------------
        # EFFECTIFS (KPI)
        # ----------------------------
        effectif_mamy = len(result["mamies"])
        effectif_gros = len(result["gros"])
        effectif_superviseur = len(result["superviseurs"])
        effectif_total = (
            effectif_mamy
            + effectif_gros
            + effectif_superviseur
        )
        
        # ----------------------------
        # CONTEXTE (ALIGNÉ AU TEMPLATE)
        # ----------------------------
        context.update({
            # période
            "date_debut": date_debut,
            "date_fin": date_fin,
            "selected_month": month,
            "selected_year": year,
            "months": [
                (1, "Janvier"), (2, "Février"), (3, "Mars"),
                (4, "Avril"), (5, "Mai"), (6, "Juin"),
                (7, "Juillet"), (8, "Août"), (9, "Septembre"),
                (10, "Octobre"), (11, "Novembre"), (12, "Décembre"),
            ],
            "years": range(today.year - 3, today.year + 1),
        
            # 🔴 LISTES
            "salaires_mamy": result["mamies"],
            "salaires_gros": result["gros"],
            "salaires_superviseur": result["superviseurs"],
        
            # 🟢 EFFECTIFS
            "effectif_mamy": effectif_mamy,
            "effectif_gros": effectif_gros,
            "effectif_superviseur": effectif_superviseur,
            "effectif_total": effectif_total,
        
            # 🟢 TOTAUX
            "total_mamy_kilo": result["total_mamy_kilo"],
            "total_mamy_salaire_base": result["total_mamy_salaire_base"],
            "total_mamy_incentive": result["total_mamy_incentive"],
            "total_mamy_general": result["total_mamy_general"],
        
            "total_gros_cartons": result["total_gros_cartons"],
            "total_gros_incentive": result["total_gros_incentive"],
            "total_gros_general": result["total_gros_general"],
        
            "total_sup_kilo_mamies": result["total_sup_kilo_mamies"],
            "total_sup_salaire_base": result["total_sup_salaire_base"],
            "total_sup_dotation": result["total_sup_dotation"],
            "total_sup_bonus": result["total_sup_bonus"],
            "total_sup_general": result["total_sup_general"],
        
            # 🌍 GLOBAL
            "total_global": result["total_global"],
        })
        
        return context


def export_salaires_mamies_excel(request):
    today = timezone.now().date()

    month = request.GET.get("month")
    year = request.GET.get("year")

    if month and year:
        month = int(month)
        year = int(year)
    else:
        month = today.month
        year = today.year

    last_day = monthrange(year, month)[1]
    date_debut = date(year, month, 1)
    date_fin = date(year, month, last_day)

    # 🔹 On filtre uniquement les mamies
    result = SalaireListeService.get_salaires(
        date_debut=date_debut,
        date_fin=date_fin,
        type_agent_filter="terrain"
    )

    mamies = result["mamies"]

    # ==========================
    # CREATION EXCEL
    # ==========================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salaires Mamies"

    headers = [
        "Agent",
        "Superviseur",
        "Kg vendus",
        "Salaire base",
        "Incentive",
        "Salaire total",
        "Jours travaillés",
        "Jours du mois",
    ]

    ws.append(headers)

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for s in mamies:
        ws.append([
            s["agent"].full_name,
            s["superviseur"].full_name if s["superviseur"] else "",
            float(s["kilo_total"]),
            float(s["salaire_base"]),
            float(s["incentive"]),
            float(s["salaire_total"]),
            s.get("jours_travailles"),
            s.get("jours_mois"),
        ])

    filename = f"salaires_mamies_{month}_{year}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


def export_salaires_gros_excel(request):
    today = timezone.now().date()

    month = request.GET.get("month")
    year = request.GET.get("year")

    if month and year:
        month = int(month)
        year = int(year)
    else:
        month = today.month
        year = today.year

    last_day = monthrange(year, month)[1]
    date_debut = date(year, month, 1)
    date_fin = date(year, month, last_day)

    # 🔸 FILTRE AGENTS GROS
    result = SalaireListeService.get_salaires(
        date_debut=date_debut,
        date_fin=date_fin,
        type_agent_filter="agent_gros"
    )

    agents = result["gros"]

    # ==========================
    # CREATION EXCEL
    # ==========================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salaires Agents Gros"

    headers = [
        "Agent",
        "Cartons vendus",
        "Incentive",
        "Salaire total",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for s in agents:
        ws.append([
            s["agent"].full_name,
            float(s["cartons_total"]),
            float(s["incentive"]),
            float(s["salaire_total"]),
        ])

    filename = f"salaires_agents_gros_{month}_{year}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


class SalaireGenerationView(LoginRequiredMixin, View):
    template_name = "paie/salaire_generation.html"

    def get(self, request):
        today = timezone.now().date()
        date_debut = today.replace(day=1)
        date_fin = today

        # Vérifier s'il y a déjà des salaires (validés ou non)
        salaires_existants = Salaire.objects.filter(
            date_debut=date_debut,
            date_fin=date_fin
        )
        
        deja_genere = salaires_existants.exists()
        deja_valide = salaires_existants.filter(valide=True).exists()

        return render(request, self.template_name, {
            "date_debut": date_debut,
            "date_fin": date_fin,
            "deja_genere": deja_genere,
            "deja_valide": deja_valide,
        })

    def post(self, request):
        try:
            date_debut = datetime.strptime(
                request.POST.get("date_debut"), "%Y-%m-%d"
            ).date()
            date_fin = datetime.strptime(
                request.POST.get("date_fin"), "%Y-%m-%d"
            ).date()

            # Vérifier s'il y a des salaires déjà validés
            salaires_valides = Salaire.objects.filter(
                date_debut=date_debut,
                date_fin=date_fin,
                valide=True
            )
            
            if salaires_valides.exists():
                messages.warning(
                    request,
                    f"{salaires_valides.count()} salaires déjà validés pour cette période. "
                    "Impossible de regénérer."
                )
                return redirect("salaire_generation")

            # Générer/Mettre à jour
            salaires = SalaireGenerationService.generate(date_debut, date_fin)
            
            messages.success(
                request,
                f"{len(salaires)} salaires générés/mis à jour avec succès."
            )
            return redirect("salaire_lecture")

        except ValueError as e:
            messages.error(request, str(e))

        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")

        return redirect("salaire_generation")
    
class SalaireValidationView(LoginRequiredMixin, View):
    template_name = "paie/salaire_validation.html"

    def get(self, request):
        date_debut = request.GET.get("date_debut")
        date_fin = request.GET.get("date_fin")

        if not date_debut or not date_fin:
            messages.error(request, "Période invalide.")
            return redirect("salaire_lecture")

        return render(request, self.template_name, {
            "date_debut": date_debut,
            "date_fin": date_fin,
        })

    def post(self, request):
        try:
            date_debut = datetime.strptime(
                request.POST.get("date_debut"), "%Y-%m-%d"
            ).date()
            date_fin = datetime.strptime(
                request.POST.get("date_fin"), "%Y-%m-%d"
            ).date()

            qs = Salaire.objects.filter(
                date_debut=date_debut,
                date_fin=date_fin,
                valide=False
            )

            if not qs.exists():
                messages.warning(request, "Aucun salaire à valider.")
                return redirect("salaire_lecture")

            qs.update(valide=True)

            messages.success(
                request,
                "Salaires validés avec succès. Ils sont maintenant figés."
            )

        except Exception as e:
            messages.error(request, f"Erreur : {e}")

        return redirect("salaire_lecture")
