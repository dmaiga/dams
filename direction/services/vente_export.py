# services/vente_export.py
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from io import BytesIO
import openpyxl
from openpyxl.styles import Font


class VenteExportService:

    # -------------------------------------------------------------------

    @staticmethod
    def export_excel(ventes_queryset, date_debut, date_fin):

        ventes_queryset = ventes_queryset.select_related("agent", "agent__user")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventes"

        headers = [
            "Date vente", "Agent", "Marché",
            "Produit", "Qté", "Prix unitaire",
            "Montant total", "Type"
        ]
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)

        row_index = 2
        for v in ventes_queryset:

            # 🔥 FIX timezone → Excel compatible
            date_naive = v.date_vente.replace(tzinfo=None)

            cell_date = ws.cell(row=row_index, column=1, value=date_naive)
            cell_date.number_format = "DD/MM/YYYY"

            ws.cell(row=row_index, column=2, value=v.agent.full_name)
            ws.cell(row=row_index, column=3, value=v.agent.marche_affectation or "Non défini")
            ws.cell(row=row_index, column=4, value=v.produit_nom)
            ws.cell(row=row_index, column=5, value=float(v.quantite))
            ws.cell(row=row_index, column=6, value=float(v.prix_vente_unitaire))
            ws.cell(row=row_index, column=7, value=float(v.total_vente))
            ws.cell(row=row_index, column=8, value=v.get_type_vente_display())

            row_index += 1

        # Auto width
        for column_cells in ws.columns:
            max_length = 0
            col = column_cells[0].column_letter

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col].width = max_length + 2

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer
    # -------------------------------------------------------------------
    @staticmethod
    def export_pdf(ventes_queryset, date_debut, date_fin):
        """Retourne un fichier PDF (BytesIO)"""

        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            title="Liste des ventes"
        )

        styles = getSampleStyleSheet()
        elements = []

        # Titre
        titre = Paragraph(
            f"<b>Liste des ventes du {date_debut.date()} au {date_fin.date()}</b>",
            styles["Title"]
        )
        elements.append(titre)

        # Tableau PDF
        data = [
            ["Date vente", "Agent", "Marche", "Produit", "Qté",
             "PU", "Montant", "Type"]
        ]

        for v in ventes_queryset:
            data.append([
                v.date_vente.strftime("%d/%m/%Y"),
                v.agent.full_name,
                v.agent.marche_affectation or "Non défini",
                v.produit_nom,
                float(v.quantite),
                float(v.prix_vente_unitaire),
                float(v.total_vente),
                v.get_type_vente_display(),
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ]))

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        return buffer
